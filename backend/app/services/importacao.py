"""Serviços para importação de extratos bancários"""
import csv
import re
from datetime import datetime
from decimal import Decimal
from xml.etree import ElementTree as ET
from typing import List, Dict, Any
from io import StringIO, BytesIO
import pdfplumber
import openpyxl


def parse_xlsx_mercado_pago(content: bytes) -> List[Dict[str, Any]]:
    """Parse XLSX específico do Mercado Pago (lê células D5 em diante)"""
    transacoes = []
    
    try:
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active
        
        print(f"Planilha ativa: {sheet.title}")
        print(f"Dimensão: {sheet.dimensions}")
        print(f"Total de linhas: {sheet.max_row}")
        print(f"Total de colunas: {sheet.max_column}")
        
        # Encontrar linha com RELEASE_DATE (cabeçalho)
        header_row = None
        print("\nProcurando cabeçalho RELEASE_DATE...")
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any('RELEASE_DATE' in str(cell) for cell in row if cell):
                header_row = row_idx
                print(f"Cabeçalho encontrado na linha {header_row}: {row}")
                break
        
        if not header_row:
            print("Cabeçalho não encontrado, usando linha 5 como fallback")
            header_row = 5
        
        print(f"\nProcessando a partir da linha {header_row + 1}")
        
        # Ler a partir da linha seguinte ao cabeçalho
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row+1, values_only=True), header_row+1):
            try:
                # Pular linhas vazias
                if not row or all(cell is None for cell in row):
                    continue
                
                print(f"\nLinha {row_idx}: {row}")
                
                # Extrair dados das colunas (A=0, B=1, C=2, D=3...)
                data_cell = row[0] if len(row) > 0 else None
                tipo_cell = row[1] if len(row) > 1 else None
                descricao_cell = row[2] if len(row) > 2 else None
                valor_cell = row[3] if len(row) > 3 else None
                
                print(f"  Data: {data_cell} | Tipo: {tipo_cell} | Desc: {descricao_cell} | Valor: {valor_cell}")
                
                if not data_cell or not valor_cell:
                    print("  -> Pulando: data ou valor vazios")
                    continue
                
                # Parse data
                data = None
                print(f"  -> Parseando data: '{data_cell}' (tipo: {type(data_cell)})")
                
                if isinstance(data_cell, datetime):
                    data = data_cell.date()
                    print(f"  -> Data datetime detectada: {data}")
                elif isinstance(data_cell, str):
                    data_str = data_cell.strip()
                    print(f"  -> Tentando formatos para: '{data_str}'")
                    
                    # Tentar diferentes formatos
                    formatos = ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y']
                    for formato in formatos:
                        try:
                            data = datetime.strptime(data_str, formato).date()
                            print(f"  -> Data convertida com formato '{formato}': {data}")
                            break
                        except:
                            print(f"  -> Formato '{formato}' falhou")
                            continue
                    
                    if not data:
                        print(f"  -> ERRO: nenhum formato funcionou para '{data_str}'")
                        continue
                
                if not data:
                    print("  -> Pulando: data não pôde ser convertida")
                    continue
                
                # Parse valor
                valor = None
                print(f"  -> Parseando valor: '{valor_cell}' (tipo: {type(valor_cell)})")
                
                if isinstance(valor_cell, (int, float)):
                    valor = Decimal(str(valor_cell))
                    print(f"  -> Valor numérico detectado: {valor}")
                elif isinstance(valor_cell, str):
                    valor_str = valor_cell.strip().replace('R$', '').replace('.', '').replace(',', '.')
                    print(f"  -> Valor string processado: '{valor_str}' -> '{valor_cell}'")
                    try:
                        valor = Decimal(valor_str)
                        print(f"  -> Valor convertido: {valor}")
                    except Exception as e:
                        print(f"  -> ERRO ao converter valor: {e}")
                        continue
                
                if valor is None:
                    print("  -> Pulando: valor não pôde ser convertido")
                    continue
                
                # Descrição
                descricao = str(tipo_cell) if tipo_cell else 'Transação Mercado Pago'
                if descricao_cell and descricao_cell != tipo_cell:
                    descricao = f"{descricao} - {descricao_cell}"
                
                # Determinar tipo (entrada/saída) - PRIORIDADE: sinal do valor
                valor_original = valor
                if valor < 0:
                    tipo = 'saida'
                    valor = abs(valor)
                    print(f"  -> Valor negativo detectado: {valor_original} -> SAÍDA {valor}")
                else:
                    tipo = 'entrada'
                    print(f"  -> Valor positivo detectado: {valor_original} -> ENTRADA {valor}")
                
                # Verificar se descrição confere (só para debug)
                desc_lower = descricao.lower()
                if tipo == 'saida' and any(x in desc_lower for x in ['recebimento', 'transferência recebida', 'rendimentos', 'estorno', 'crédito', 'deposito', 'dinheiro reservado']):
                    print(f"  -> ALERTA: Descrição sugere entrada mas valor é negativo: {descricao}")
                elif tipo == 'entrada' and any(x in desc_lower for x in ['pagamento', 'envio', 'transferência enviada', 'dinheiro retirado', 'pagamento com qr', 'compra', 'debito']):
                    print(f"  -> ALERTA: Descrição sugere saída mas valor é positivo: {descricao}")
                
                transacoes.append({
                    'descricao': descricao.strip(),
                    'valor': float(valor),
                    'tipo': tipo,
                    'data': data.isoformat(),
                    'categoria': detectar_categoria(descricao),
                    'is_fixa': False,
                    'is_recorrente': 'rendimento' in desc_lower,
                    'observacoes': 'Importado via XLSX Mercado Pago'
                })
                
            except Exception as e:
                print(f"Erro ao processar linha {row_idx}: {e}")
                continue
                
    except Exception as e:
        print(f"Erro ao processar XLSX: {e}")
    
    return transacoes


def parse_mercado_pago_csv(content: str) -> List[Dict[str, Any]]:
    """Parse CSV específico do Mercado Pago (formato com RELEASE_DATE, TRANSACTION_TYPE, etc.)"""
    transacoes = []
    lines = content.strip().split('\n')
    
    # Remover linhas de saldo inicial/final e cabeçalho
    header_idx = None
    for i, line in enumerate(lines):
        if 'RELEASE_DATE' in line.upper() or 'TRANSACTION_TYPE' in line.upper():
            header_idx = i
            break
    
    if header_idx is None:
        # Tentar detectar formato alternativo
        return parse_csv(content, 'mercado_pago')
    
    # Processar a partir do cabeçalho
    reader = csv.DictReader(lines[header_idx:])
    
    for row in reader:
        try:
            # Mapear campos do Mercado Pago
            data_str = row.get('RELEASE_DATE', '')
            tipo_transacao = row.get('TRANSACTION_TYPE', '').strip()
            descricao = row.get('DESCRIPTION', '') or row.get('TRANSACTION_TYPE', '')
            valor_str = row.get('TRANSACTION_NET_AMOUNT', '') or row.get('NET_AMOUNT', '')
            
            if not data_str or not valor_str:
                continue
            
            # Parse data (formato DD/MM/YYYY do Mercado Pago)
            data = None
            try:
                data = datetime.strptime(data_str.strip(), '%d/%m/%Y').date()
            except:
                try:
                    data = datetime.strptime(data_str.strip(), '%Y-%m-%d').date()
                except:
                    pass
            
            if not data:
                continue
            
            # Parse valor (formato brasileiro com vírgula)
            valor_str = valor_str.strip().replace('R$', '').replace('.', '').replace(',', '.')
            try:
                valor = Decimal(valor_str)
            except:
                continue
            
            # Determinar tipo baseado no valor e no tipo de transação
            tipo_transacao_lower = tipo_transacao.lower()
            
            # Transações de saída comuns no Mercado Pago
            saidas = ['pagamento', 'envio', 'transferência enviada', 'dinheiro retirado', 
                      'pagamento com qr', 'compra', 'debito']
            # Transações de entrada
            entradas = ['recebimento', 'transferência recebida', 'rendimentos', 'estorno',
                        'crédito', 'deposito', 'dinheiro reservado']
            
            if any(s in tipo_transacao_lower for s in saidas):
                tipo = 'saida'
                valor = abs(valor)
            elif any(e in tipo_transacao_lower for e in entradas):
                tipo = 'entrada'
                valor = abs(valor)
            else:
                # Determinar pelo sinal do valor
                if valor < 0:
                    tipo = 'saida'
                    valor = abs(valor)
                else:
                    tipo = 'entrada'
            
            transacoes.append({
                'descricao': descricao.strip() or tipo_transacao,
                'valor': float(valor),
                'tipo': tipo,
                'data': data.isoformat(),
                'categoria': detectar_categoria(descricao or tipo_transacao),
                'is_fixa': False,
                'is_recorrente': 'rendimento' in tipo_transacao_lower,
                'observacoes': f'Mercado Pago - {tipo_transacao}'
            })
            
        except Exception as e:
            print(f"Erro ao parsear linha Mercado Pago: {e}")
            continue
    
    return transacoes


def parse_pdf_mercado_pago(content: bytes) -> List[Dict[str, Any]]:
    """Parse PDF do Mercado Pago e extrai transações"""
    transacoes = []
    
    try:
        with pdfplumber.open(BytesIO(content)) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
            
            # Padrões de linha de transação no PDF do Mercado Pago
            # Formato: DATA DESCRIÇÃO ID_OPERAÇÃO VALOR SALDO
            linhas = texto_completo.split('\n')
            
            for linha in linhas:
                try:
                    # Tentar extrair data (DD-MM-YYYY ou DD/MM/YYYY)
                    match_data = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', linha)
                    if not match_data:
                        continue
                    
                    dia, mes, ano = match_data.groups()
                    data = datetime(int(ano), int(mes), int(dia)).date()
                    
                    # Tentar extrair valor (pegar apenas o PRIMEIRO R$ - valor da transação)
                    match_valores = re.findall(r'R\$\s*([\d.,-]+)', linha)
                    if not match_valores:
                        continue
                    
                    # Usar apenas o primeiro valor (transação), ignorar o saldo
                    valor_str = match_valores[0].replace('.', '').replace(',', '.')
                    try:
                        valor = Decimal(valor_str)
                    except:
                        continue
                    
                    # Extrair descrição (tudo entre data e o primeiro valor)
                    primeiro_valor_pos = linha.find(f'R$ {match_valores[0]}')
                    descricao = linha[match_data.end():primeiro_valor_pos].strip()
                    
                    # Limpar ID de operação se presente
                    descricao = re.sub(r'\d{12,}', '', descricao).strip()
                    
                    if not descricao:
                        continue
                    
                    # Determinar tipo
                    desc_lower = descricao.lower()
                    if any(x in desc_lower for x in ['pagamento', 'retirado', 'enviado']):
                        tipo = 'saida'
                        valor = abs(valor)
                    elif any(x in desc_lower for x in ['recebido', 'rendimentos', 'reservado']):
                        tipo = 'entrada'
                        valor = abs(valor)
                    else:
                        tipo = 'saida' if valor < 0 else 'entrada'
                        valor = abs(valor)
                    
                    transacoes.append({
                        'descricao': descricao,
                        'valor': float(valor),
                        'tipo': tipo,
                        'data': data.isoformat(),
                        'categoria': detectar_categoria(descricao),
                        'is_fixa': False,
                        'is_recorrente': 'rendimento' in desc_lower,
                        'observacoes': 'Importado via PDF Mercado Pago'
                    })
                    
                except Exception as e:
                    continue
                    
    except Exception as e:
        print(f"Erro ao processar PDF: {e}")
    
    return transacoes


def parse_ofx(content: str) -> List[Dict[str, Any]]:
    """Parse arquivo OFX e retorna lista de transações"""
    transacoes = []
    
    # Extrair transações do OFX (formato SGML/XML)
    # OFX é um formato híbrido, vamos tratar como SGML
    
    # Padrão para transações no OFX
    stmttrn_pattern = re.findall(
        r'<STMTTRN>(.*?)</STMTTRN>',
        content,
        re.DOTALL | re.IGNORECASE
    )
    
    for trn in stmttrn_pattern:
        try:
            # Extrair tipo (DEBIT/CREDIT)
            trntype = re.search(r'<TRNTYPE>(.*?)</TRNTYPE>', trn, re.IGNORECASE)
            trntype = trntype.group(1).upper() if trntype else 'OTHER'
            
            # Extrair data
            dtposted = re.search(r'<DTPOSTED>(.*?)</DTPOSTED>', trn, re.IGNORECASE)
            if dtposted:
                data_str = dtposted.group(1)[:8]  # YYYYMMDD
                data = datetime.strptime(data_str, '%Y%m%d').date()
            else:
                data = datetime.now().date()
            
            # Extrair valor
            trnamt = re.search(r'<TRNAMT>(.*?)</TRNAMT>', trn, re.IGNORECASE)
            if trnamt:
                valor = Decimal(trnamt.group(1).replace(',', ''))
            else:
                continue
            
            # Extrair descrição/memo
            memo = re.search(r'<MEMO>(.*?)</MEMO>', trn, re.IGNORECASE)
            name = re.search(r'<NAME>(.*?)</NAME>', trn, re.IGNORECASE)
            
            descricao = ''
            if memo:
                descricao = memo.group(1)
            elif name:
                descricao = name.group(1)
            else:
                descricao = 'Transação bancária'
            
            # Determinar tipo (entrada/saída)
            if trntype == 'CREDIT' or valor > 0:
                tipo = 'entrada'
            elif trntype == 'DEBIT' or valor < 0:
                tipo = 'saida'
                valor = abs(valor)
            else:
                tipo = 'saida'
            
            transacoes.append({
                'descricao': descricao.strip(),
                'valor': float(valor),
                'tipo': tipo,
                'data': data.isoformat(),
                'categoria': detectar_categoria(descricao),
                'is_fixa': False,
                'is_recorrente': False,
                'observacoes': f'Importado via OFX - Tipo: {trntype}'
            })
            
        except Exception as e:
            print(f"Erro ao parsear transação OFX: {e}")
            continue
    
    return transacoes


def parse_csv(content: str, banco: str = 'auto') -> List[Dict[str, Any]]:
    """Parse arquivo CSV e retorna lista de transações"""
    transacoes = []
    
    # Detectar delimitador
    if ';' in content[:1000]:
        delimiter = ';'
    else:
        delimiter = ','
    
    # Tentar detectar banco pelo conteúdo
    if banco == 'auto':
        banco = detectar_banco_csv(content)
    
    try:
        reader = csv.DictReader(StringIO(content), delimiter=delimiter)
        
        for row in reader:
            try:
                transacao = extrair_transacao_csv(row, banco)
                if transacao:
                    transacoes.append(transacao)
            except Exception as e:
                print(f"Erro ao processar linha CSV: {e}")
                continue
                
    except Exception as e:
        # Tentar com diferentes configurações
        try:
            lines = content.strip().split('\n')
            for line in lines[1:]:  # Pular header
                cols = line.split(delimiter)
                if len(cols) >= 3:
                    transacao = extrair_transacao_csv_generico(cols)
                    if transacao:
                        transacoes.append(transacao)
        except Exception as e2:
            print(f"Erro ao parsear CSV: {e2}")
    
    return transacoes


def detectar_banco_csv(content: str) -> str:
    """Detecta qual banco gerou o CSV baseado no conteúdo"""
    content_upper = content.upper()
    
    if 'MERCADO PAGO' in content_upper or 'RELEASE_DATE' in content_upper or 'TRANSACTION_TYPE' in content_upper:
        return 'mercado_pago'
    elif 'NU PAGAMENTOS' in content_upper or 'NU PAGAMENTO' in content_upper or 'NUBANK' in content_upper:
        return 'nubank'
    elif 'ITAU' in content_upper or 'ITAÚ' in content_upper:
        return 'itau'
    elif 'BRADESCO' in content_upper:
        return 'bradesco'
    elif 'SANTANDER' in content_upper:
        return 'santander'
    elif 'BB' in content_upper or 'BANCO DO BRASIL' in content_upper:
        return 'bb'
    elif 'CAIXA' in content_upper:
        return 'caixa'
    elif 'INTER' in content_upper:
        return 'inter'
    elif 'C6' in content_upper:
        return 'c6'
    elif 'ORIGINAL' in content_upper:
        return 'original'
    else:
        return 'generico'


def extrair_transacao_csv(row: Dict[str, str], banco: str) -> Dict[str, Any]:
    """Extrai transação de uma linha CSV baseado no banco"""
    
    # Mapeamento de campos por banco
    mapeamentos = {
        'nubank': {
            'data': ['Data', 'data', 'DATA'],
            'descricao': ['Descrição', 'descricao', 'DESCRICAO', 'Estabelecimento'],
            'valor': ['Valor', 'valor', 'VALOR', 'Valor (R$)'],
            'tipo': None  # Determinar pelo sinal do valor
        },
        'itau': {
            'data': ['Data', 'data', 'DATA', 'Data Lançamento'],
            'descricao': ['Histórico', 'historico', 'HISTORICO', 'Lançamento', 'Descrição'],
            'valor': ['Valor', 'valor', 'VALOR', 'Valor (R$)', 'Valor R$'],
            'tipo': ['Tipo', 'tipo', 'TIPO', 'Entrada/Saída']
        },
        'generico': {
            'data': ['Data', 'data', 'DATA', 'Date', 'DATE', 'Dt', 'DT'],
            'descricao': ['Descricao', 'descricao', 'DESCRICAO', 'Descrição', 'Historico', 'Lancamento', 'Movimentação', 'Movimentacao'],
            'valor': ['Valor', 'valor', 'VALOR', 'Value', 'Amount', 'Valor R$', 'Valor (R$)'],
            'tipo': ['Tipo', 'tipo', 'TIPO', 'Entrada/Saída', 'Natureza']
        }
    }
    
    mapeamento = mapeamentos.get(banco, mapeamentos['generico'])
    
    # Encontrar campos
    data_str = encontrar_campo(row, mapeamento['data'])
    descricao = encontrar_campo(row, mapeamento['descricao'])
    valor_str = encontrar_campo(row, mapeamento['valor'])
    tipo_str = encontrar_campo(row, mapeamento['tipo']) if mapeamento['tipo'] else None
    
    if not all([data_str, descricao, valor_str]):
        return None
    
    # Parse data
    data = parse_data(data_str)
    
    # Parse valor
    valor = parse_valor(valor_str)
    
    # Determinar tipo
    if tipo_str:
        tipo_str_lower = tipo_str.lower()
        if 'entrada' in tipo_str_lower or 'credito' in tipo_str_lower or 'receita' in tipo_str_lower:
            tipo = 'entrada'
        elif 'saida' in tipo_str_lower or 'debito' in tipo_str_lower or 'despesa' in tipo_str_lower:
            tipo = 'saida'
        else:
            tipo = 'saida' if valor < 0 else 'entrada'
    else:
        # Determinar pelo sinal do valor
        if valor < 0:
            tipo = 'saida'
            valor = abs(valor)
        else:
            tipo = 'entrada'
    
    return {
        'descricao': descricao.strip(),
        'valor': float(valor),
        'tipo': tipo,
        'data': data.isoformat() if data else datetime.now().date().isoformat(),
        'categoria': detectar_categoria(descricao),
        'is_fixa': False,
        'is_recorrente': False,
        'observacoes': f'Importado via CSV - Banco: {banco}'
    }


def extrair_transacao_csv_generico(cols: List[str]) -> Dict[str, Any]:
    """Extrai transação de colunas genéricas"""
    if len(cols) < 3:
        return None
    
    # Tentar identificar qual coluna é data, descrição e valor
    data_idx = None
    valor_idx = None
    desc_idx = None
    
    for i, col in enumerate(cols):
        col_upper = col.upper().strip()
        
        # Detectar data
        if data_idx is None and any(x in col_upper for x in ['DATA', 'DATE', 'DT']):
            data_idx = i
        
        # Detectar valor
        if valor_idx is None and any(x in col_upper for x in ['VALOR', 'VALUE', 'AMOUNT', 'R$']):
            valor_idx = i
        
        # Detectar descrição
        if desc_idx is None and any(x in col_upper for x in ['DESC', 'HIST', 'LANC', 'MOVIM']):
            desc_idx = i
    
    # Se não detectou, usar posições padrão
    if data_idx is None:
        data_idx = 0
    if desc_idx is None:
        desc_idx = 1 if len(cols) > 1 else 0
    if valor_idx is None:
        valor_idx = len(cols) - 1
    
    try:
        data = parse_data(cols[data_idx])
        descricao = cols[desc_idx]
        valor = parse_valor(cols[valor_idx])
        
        tipo = 'saida' if valor < 0 else 'entrada'
        if valor < 0:
            valor = abs(valor)
        
        return {
            'descricao': descricao.strip(),
            'valor': float(valor),
            'tipo': tipo,
            'data': data.isoformat() if data else datetime.now().date().isoformat(),
            'categoria': detectar_categoria(descricao),
            'is_fixa': False,
            'is_recorrente': False,
            'observacoes': 'Importado via CSV'
        }
    except:
        return None


def encontrar_campo(row: Dict[str, str], possiveis_nomes: List[str]) -> str:
    """Encontra valor do campo baseado em possíveis nomes"""
    if not possiveis_nomes:
        return None
    
    for nome in possiveis_nomes:
        if nome in row:
            return row[nome]
    
    # Tentar case-insensitive
    row_lower = {k.lower(): v for k, v in row.items()}
    for nome in possiveis_nomes:
        if nome.lower() in row_lower:
            return row_lower[nome.lower()]
    
    return None


def parse_data(data_str: str) -> datetime.date:
    """Parse string de data em vários formatos"""
    formatos = [
        '%d/%m/%Y',
        '%d/%m/%y',
        '%Y-%m-%d',
        '%Y%m%d',
        '%d-%m-%Y',
        '%d-%m-%y',
        '%m/%d/%Y',
        '%d.%m.%Y',
    ]
    
    data_str = data_str.strip()
    
    for formato in formatos:
        try:
            return datetime.strptime(data_str, formato).date()
        except ValueError:
            continue
    
    # Tentar detectar padrão DD/MM/AAAA ou similar
    if re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', data_str):
        partes = re.split(r'[/-]', data_str)
        if len(partes) == 3:
            try:
                dia, mes, ano = int(partes[0]), int(partes[1]), int(partes[2])
                if ano < 100:
                    ano += 2000 if ano < 50 else 1900
                return datetime(ano, mes, dia).date()
            except:
                pass
    
    return None


def parse_valor(valor_str: str) -> Decimal:
    """Parse string de valor em Decimal"""
    if not valor_str:
        return Decimal('0')
    
    # Limpar string
    valor_str = valor_str.strip()
    
    # Remover símbolo de moeda
    valor_str = re.sub(r'R\$|\$', '', valor_str)
    
    # Detectar formato
    # Brasileiro: 1.234,56 ou 1234,56
    # Americano: 1,234.56 ou 1234.56
    
    if ',' in valor_str and '.' in valor_str:
        # Ambos presentes - detectar qual é decimal
        last_comma = valor_str.rfind(',')
        last_dot = valor_str.rfind('.')
        
        if last_comma > last_dot:
            # Formato brasileiro: 1.234,56
            valor_str = valor_str.replace('.', '').replace(',', '.')
        else:
            # Formato americano: 1,234.56
            valor_str = valor_str.replace(',', '')
    elif ',' in valor_str:
        # Pode ser 1234,56 (brasileiro) ou 1,234 (americano)
        # Se tiver 2 ou menos dígitos após vírgula, é decimal
        parts = valor_str.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            valor_str = valor_str.replace(',', '.')
        else:
            valor_str = valor_str.replace(',', '')
    
    # Remover espaços
    valor_str = valor_str.replace(' ', '')
    
    return Decimal(valor_str)


def detectar_categoria(descricao: str) -> str:
    """Detecta categoria baseada na descrição"""
    desc_lower = descricao.lower()
    
    categorias = {
        'alimentacao': ['mercado', 'supermercado', 'restaurante', 'lanchonete', 'padaria', 'ifood', 'uber eats', 'rappi', 'comida', 'alimentação'],
        'transporte': ['uber', '99', 'taxi', 'ônibus', 'onibus', 'combustivel', 'gasolina', 'posto', 'estacionamento', 'pedagio', 'parking'],
        'moradia': ['aluguel', 'condominio', 'condomínio', 'iptu', 'luz', 'água', 'agua', 'gas', 'gás', 'internet', 'telefone', 'eletricidade'],
        'saude': ['farmacia', 'farmacias', 'hospital', 'clinica', 'consulta', 'medico', 'plano de saude', 'dentista'],
        'educacao': ['escola', 'faculdade', 'universidade', 'curso', 'material escolar', 'mensalidade', 'livraria'],
        'lazer': ['cinema', 'teatro', 'show', 'bar', 'balada', 'viagem', 'hotel', 'passagem', 'streaming', 'netflix', 'spotify', 'youtube', 'amazon prime'],
        'vestuario': ['loja', 'roupa', 'camisa', 'calca', 'calça', 'tenis', 'sapato', 'shopping', 'magazine', 'shein', 'amazon'],
        'utilidades': ['celular', 'telefone', 'assinatura', 'servicos', 'manutencao', 'reparo'],
        'salario': ['salario', 'pagamento', 'prolabore', 'recebimento', 'transferencia recebida'],
        'investimentos': ['investimento', 'aplicacao', 'cdb', 'tesouro', 'acoes', 'ações', 'fundo', 'rendimento'],
    }
    
    for categoria, palavras in categorias.items():
        for palavra in palavras:
            if palavra in desc_lower:
                return categoria
    
    return 'outros'


def importar_transacoes(content, formato: str, banco: str = 'auto') -> List[Dict[str, Any]]:
    """Importa transações de arquivo"""
    if formato.lower() == 'ofx':
        return parse_ofx(content)
    elif formato.lower() == 'csv':
        # Detectar se é formato específico do Mercado Pago
        if 'mercado_pago' in banco.lower() or 'RELEASE_DATE' in content[:2000].upper():
            return parse_mercado_pago_csv(content)
        return parse_csv(content, banco)
    elif formato.lower() == 'xlsx':
        # XLSX geralmente é do Mercado Pago
        return parse_xlsx_mercado_pago(content)
    elif formato.lower() == 'pdf':
        return parse_pdf_mercado_pago(content.encode() if isinstance(content, str) else content)
    else:
        # Tentar detectar formato
        if isinstance(content, bytes):
            # Pode ser XLSX ou PDF
            try:
                # Tentar como XLSX primeiro
                return parse_xlsx_mercado_pago(content)
            except:
                # Tentar como PDF
                return parse_pdf_mercado_pago(content)
        else:
            content_upper = content[:2000].upper()
            if '<OFX>' in content_upper or '<STMTTRN>' in content_upper:
                return parse_ofx(content)
            elif 'RELEASE_DATE' in content_upper or 'TRANSACTION_TYPE' in content_upper:
                return parse_mercado_pago_csv(content)
            else:
                return parse_csv(content, banco)
